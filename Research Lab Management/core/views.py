
from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook
from .models import Project, Publication, Grant, Equipment, Inventory, Staff

def index(request):
    return render(request, 'core/RLM.html')

def export_report(request):
    try:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="research_report.xlsx"'

        workbook = Workbook()
        
        def safe_str(value):
            return str(value) if value is not None else 'N/A'

        # Projects worksheet
        try:
            projects_sheet = workbook.active
            projects_sheet.title = 'Projects'
            projects_sheet.append(['Name', 'Lead Researcher', 'Status', 'Start Date', 'End Date'])
            for project in Project.objects.all():
                projects_sheet.append([
                    safe_str(project.name),
                    safe_str(project.lead_researcher.username),
                    safe_str(project.get_status_display()),
                    safe_str(project.start_date),
                    safe_str(project.end_date),
                ])
        except Exception as e:
            projects_sheet.append(['Error loading projects data:', str(e)])

        # Equipment worksheet
        try:
            equipment_sheet = workbook.create_sheet(title='Equipment')
            equipment_sheet.append(['Name', 'Type', 'Status', 'Location'])
            for equipment in Equipment.objects.all():
                equipment_sheet.append([
                    safe_str(equipment.name),
                    safe_str(equipment.type),
                    safe_str(equipment.status),
                    safe_str(equipment.location),
                ])
        except Exception as e:
            equipment_sheet.append(['Error loading equipment data:', str(e)])

        # Inventory worksheet
        try:
            inventory_sheet = workbook.create_sheet(title='Inventory')
            inventory_sheet.append(['Item Name', 'Quantity', 'Unit', 'Location', 'CAS Number', 'Storage Conditions'])
            for item in Inventory.objects.all():
                inventory_sheet.append([
                    safe_str(item.item_name),
                    safe_str(item.quantity),
                    safe_str(item.unit),
                    safe_str(item.location),
                    safe_str(item.cas_number),
                    safe_str(item.storage_conditions),
                ])
        except Exception as e:
            inventory_sheet.append(['Error loading inventory data:', str(e)])

        # Staff worksheet
        try:
            staff_sheet = workbook.create_sheet(title='Staff')
            staff_sheet.append(['Name', 'Role', 'Email', 'Phone'])
            for staff in Staff.objects.all():
                staff_sheet.append([
                    safe_str(staff.name),
                    safe_str(staff.role),
                    safe_str(staff.email),
                    safe_str(staff.phone),
                ])
        except Exception as e:
            staff_sheet.append(['Error loading staff data:', str(e)])

        # Publications worksheet
        try:
            publications_sheet = workbook.create_sheet(title='Publications')
            publications_sheet.append(['Title', 'Authors', 'Journal/Conference', 'Year'])
            for publication in Publication.objects.all():
                publications_sheet.append([
                    safe_str(publication.title),
                    safe_str(publication.authors),
                    safe_str(publication.journal or publication.conference),
                    safe_str(publication.year),
                ])
        except Exception as e:
            publications_sheet.append(['Error loading publications data:', str(e)])

        # Grants worksheet
        try:
            grants_sheet = workbook.create_sheet(title='Grants')
            grants_sheet.append(['Name', 'Funding Agency', 'Amount', 'Start Date', 'End Date'])
            for grant in Grant.objects.all():
                grants_sheet.append([
                    safe_str(grant.name),
                    safe_str(grant.funding_agency),
                    safe_str(grant.amount),
                    safe_str(grant.start_date),
                    safe_str(grant.end_date),
                ])
        except Exception as e:
            grants_sheet.append(['Error loading grants data:', str(e)])

        workbook.save(response)
        return response
    except Exception as e:
        return HttpResponse(f"Error generating report: {str(e)}", status=500)
